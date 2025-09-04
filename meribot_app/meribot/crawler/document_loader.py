# ================= Ejemplo de integración tras el chunking =================
# Supón que tienes un documento procesado y quieres fragmentarlo y clasificar los chunks:
#
# from meribot.crawler.document_loader import chunk_text_with_langchain, process_and_classify_chunks
#
# text = "...texto extraído del documento..."
# metadata = {"id": "doc1", "url": "http://ejemplo.com/doc1"}  # Ajusta según tus metadatos
# chunks = chunk_text_with_langchain(text)  # Usa RecursiveCharacterTextSplitter
# metadata_list = [{**metadata, "chunk_idx": i} for i in range(len(chunks))]
# resultados = process_and_classify_chunks(chunks, metadata_list)
# for chunk, meta, estado in resultados:
#     print(f"Chunk {meta['chunk_idx']} ({meta.get('id') or meta.get('url')}): {estado}")
# ===========================================================================
from dotenv import load_dotenv
load_dotenv()


import numpy as np
from typing import List, Optional
import requests
import os
import json

# --- LangChain Text Splitter ---
from langchain_text_splitters import RecursiveCharacterTextSplitter

# --- Utilidades de hash y persistencia ---
from meribot.utils.hash_utils import calculate_sha256, load_hash_db, save_hash_db

# --- Configuración del splitter ---
SPLITTER_CHUNK_SIZE_DOC = 2000
SPLITTER_CHUNK_OVERLAP_DOC = 50

# Ruta por defecto para la base de datos de hashes (puedes cambiarla)
HASH_DB_PATH = os.getenv('HASH_DB_PATH', 'hash_db.json')
# Ejemplo de función para procesar y clasificar chunks según hash
def process_and_classify_chunks(chunks: list, metadata_list: list, hash_db_path: str = HASH_DB_PATH):
    """
    Procesa una lista de chunks y los clasifica como 'nuevo', 'modificado' o 'sin cambios' usando hashes persistentes.
    :param chunks: Lista de textos (chunks)
    :param metadata_list: Lista de metadatos asociados a cada chunk (debe tener un campo 'id' o similar)
    :param hash_db_path: Ruta al archivo JSON de hashes
    :return: Lista de tuplas (chunk, metadata, estado)
    """
    hash_db = load_hash_db(hash_db_path)
    resultados = []
    for chunk, meta in zip(chunks, metadata_list):
        chunk_id = meta.get('id') or meta.get('url') or meta.get('title') or str(hash(chunk))
        chunk_hash = calculate_sha256(chunk)
        prev_hash = hash_db.get(chunk_id)
        if prev_hash is None:
            estado = 'nuevo'
        elif prev_hash == chunk_hash:
            estado = 'sin cambios'
        else:
            estado = 'modificado'
        hash_db[chunk_id] = chunk_hash
        resultados.append((chunk, meta, estado))
    save_hash_db(hash_db, hash_db_path)
    return resultados

def get_azure_openai_embeddings(sentences: list[str]) -> np.ndarray:
    """
    Obtiene embeddings de Azure OpenAI para una lista de frases.
    Requiere las variables de entorno:
        AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_KEY, AZURE_OPENAI_DEPLOYMENT
    """
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
    api_key = os.getenv("AZURE_OPENAI_KEY")
    deployment = os.getenv("AZURE_OPENAI_EMBEDDINGS_DEPLOYMENT")
    if not endpoint or not api_key or not deployment:
        raise RuntimeError("Faltan variables de entorno para Azure OpenAI embeddings.")
    url = f"{endpoint}/openai/deployments/{deployment}/embeddings?api-version=2023-05-15"
    headers = {
        "Content-Type": "application/json",
        "api-key": api_key
    }
    data = {"input": sentences}
    response = requests.post(url, headers=headers, data=json.dumps(data))
    if response.status_code != 200:
        raise RuntimeError(f"Error Azure OpenAI: {response.status_code} {response.text}")
    result = response.json()
    # El resultado es una lista de dicts con 'embedding' en el mismo orden
    embeddings = [item["embedding"] for item in result["data"]]
    return np.array(embeddings)

def chunk_text_with_langchain(
    text: str,
    chunk_size: int = SPLITTER_CHUNK_SIZE_DOC,
    chunk_overlap: int = SPLITTER_CHUNK_OVERLAP_DOC
) -> List[str]:
    """
    Divide el texto en chunks usando RecursiveCharacterTextSplitter de LangChain.
    
    Args:
        text (str): Texto a fragmentar.
        chunk_size (int): Tamaño máximo de cada chunk en caracteres.
        chunk_overlap (int): Superposición entre chunks en caracteres.
        
    Returns:
        List[str]: Lista de chunks de texto.
    """
    if not text or not text.strip():
        print("[ADVERTENCIA] Texto vacío, no se pueden crear chunks.")
        return []
    
    # Crear el splitter con los parámetros especificados
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        length_function=len,
        is_separator_regex=False,
    )
    
    # Dividir el texto en chunks
    chunks = text_splitter.split_text(text)
    
    print(f"[DEBUG] Texto dividido en {len(chunks)} chunks usando RecursiveCharacterTextSplitter")
    print(f"[DEBUG] Configuración: chunk_size={chunk_size}, chunk_overlap={chunk_overlap}")
    
    # Mostrar información de los primeros chunks para debug
    for i, chunk in enumerate(chunks[:3]):
        print(f"  Chunk {i}: {len(chunk)} chars | {chunk[:80]}{'...' if len(chunk) > 80 else ''}")
    
    if len(chunks) > 3:
        print(f"  ... y {len(chunks)-3} chunks más")
    
    return chunks

# Mantener la función original para compatibilidad hacia atrás
def semantic_chunk_text(
    text: str,
    max_chunk_size: int = 2000,
    min_chunk_size: int = 200,
    stride: int = 50,
    split_on_newline: bool = True,
    similarity_threshold: float = 0.7,
    return_embeddings: bool = False
) -> List[str]:
    """
    DEPRECATED: Usa chunk_text_with_langchain() para mejor rendimiento.
    
    Divide el texto usando RecursiveCharacterTextSplitter en lugar del chunking semántico original.
    Mantenida para compatibilidad, pero redirige a la nueva implementación.

    Args:
        text (str): Texto a fragmentar.
        max_chunk_size (int): Máximo de caracteres por chunk (usado como chunk_size).
        stride (int): Superposición de caracteres entre chunks (usado como chunk_overlap).
        Otros parámetros se ignoran para simplificar.

    Returns:
        List[str]: Lista de chunks de texto.
    """
    print("[WARNING] semantic_chunk_text() está deprecada. Usa chunk_text_with_langchain() en su lugar.")
    print(f"[DEBUG] Redirigiendo a RecursiveCharacterTextSplitter con chunk_size={max_chunk_size}, overlap={stride}")
    
    return chunk_text_with_langchain(text, chunk_size=max_chunk_size, chunk_overlap=stride)
# Fragmentación de texto
def split_text(text: str, max_length: int = 1000) -> list:
    """
    Fragmenta el texto en trozos de longitud máxima, procurando no cortar frases ni palabras.
    """
    if not isinstance(text, str) or max_length <= 0:
        return [text]
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks = []
    current = ''
    for sentence in sentences:
        if len(current) + len(sentence) + 1 <= max_length:
            current += (' ' if current else '') + sentence
        else:
            if current:
                chunks.append(current.strip())
            current = sentence
    if current:
        chunks.append(current.strip())
    return chunks
# Normalización de texto
import re

def normalize_text(text: str) -> str:
    """
    Limpia el texto eliminando espacios duplicados, saltos de línea innecesarios y caracteres de control.
    """
    if not isinstance(text, str):
        return text
    # Reemplaza saltos de línea múltiples por uno solo
    text = re.sub(r'\n+', '\n', text)
    # Reemplaza espacios múltiples por uno solo
    text = re.sub(r'[ \t]+', ' ', text)
    # Elimina caracteres de control excepto tabulador y salto de línea
    text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f]', '', text)
    # Quita espacios al inicio y final
    return text.strip()
"""
Módulo para cargar y extraer texto/metadatos de documentos soportados por el crawler.
- HTML: BeautifulSoup
- DOCX: python-docx
- XLSX: openpyxl
- PDF: PyMuPDF (fitz)
"""
import os
from typing import Dict, Any

# HTML
from bs4 import BeautifulSoup

def parse_html(content: str, url: str = None) -> Dict[str, Any]:
    """
    Extrae texto y metadatos básicos de HTML.
    :param content: HTML como string
    :return: dict con 'text' y 'metadata'
    """
    try:
        soup = BeautifulSoup(content, "html.parser")
        # Heurística: priorizar main, luego article, luego section, luego body, luego bloque más largo
        candidates = []
        for tag in ["main", "article", "section"]:
            el = soup.find(tag)
            if el and el.get_text(strip=True):
                candidates.append(el.get_text(separator=" ", strip=True))
        if not candidates and soup.body:
            candidates.append(soup.body.get_text(separator=" ", strip=True))
        # Si no hay nada, usar el bloque de texto más largo
        if not candidates:
            blocks = [b.get_text(separator=" ", strip=True) for b in soup.find_all(True) if b.get_text(strip=True)]
            if blocks:
                candidates.append(max(blocks, key=len))
        text = candidates[0] if candidates else ""
        title = soup.title.string if soup.title else None
        metadata = {
            "title": title,
            "author": None,
            "date": None,
            "version": None,
            "type": "html",
            "url": url,
            "domain": None
        }
        return {
            "text": text,
            "metadata": metadata
        }
    except Exception as e:
        return {"error": str(e), "text": None, "metadata": None}

# DOCX
try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

def parse_docx(path: str, url: str = None) -> Dict[str, Any]:
    """
    Extrae texto de un archivo Word (.docx).
    :param path: ruta al archivo docx
    :return: dict con 'text' y 'metadata'
    """
    if not DOCX_AVAILABLE:
        return {"error": "python-docx no está instalado", "text": None, "metadata": None}
    try:
        doc = docx.Document(path)
        text = "\n".join([p.text for p in doc.paragraphs])
        props = doc.core_properties
        metadata = {
            "title": getattr(props, "title", None),
            "author": getattr(props, "author", None),
            "date": getattr(props, "created", None),
            "version": getattr(props, "version", None) if hasattr(props, "version") else None,
            "type": "docx",
            "url": url,
            "domain": None
        }
        return {"text": text, "metadata": metadata}
    except Exception as e:
        return {"error": str(e), "text": None, "metadata": None}

# XLSX
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

def parse_xlsx(path: str, url: str = None) -> Dict[str, Any]:
    """
    Extrae texto de un archivo Excel (.xlsx).
    :param path: ruta al archivo xlsx
    :return: dict con 'text' y 'metadata'
    """
    if not XLSX_AVAILABLE:
        return {"error": "openpyxl no está instalado", "text": None, "metadata": None}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        text = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                text.append("\t".join([str(cell) if cell is not None else '' for cell in row]))
        metadata = {
            "title": None,
            "author": None,
            "date": None,
            "version": None,
            "type": "xlsx",
            "url": url,
            "domain": None,
            "sheets": wb.sheetnames
        }
        return {"text": "\n".join(text), "metadata": metadata}
    except Exception as e:
        return {"error": str(e), "text": None, "metadata": None}

# PDF
try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

def parse_pdf(path: str, url: str = None) -> Dict[str, Any]:
    """
    Extrae texto de un archivo PDF usando PyMuPDF.
    :param path: ruta al archivo PDF
    :return: dict con 'text' y 'metadata'
    """
    if not PDF_AVAILABLE:
        return {"error": "PyMuPDF (fitz) no está instalado", "text": None, "metadata": None}
    try:
        doc = fitz.open(path)
        text = "\n".join(page.get_text() for page in doc)
        metadata = doc.metadata
        meta = {
            "title": metadata.get("title"),
            "author": metadata.get("author"),
            "date": metadata.get("creationDate"),
            "version": metadata.get("version", None),
            "type": "pdf",
            "url": url,
            "domain": None
        }
        return {"text": text, "metadata": meta}
    except Exception as e:
        return {"error": str(e), "text": None, "metadata": None}
# Función general para seleccionar parser según extensión
def parse_document(path: str, url: str = None) -> dict:
    """
    Selecciona el parser adecuado según la extensión del archivo.
    :param path: ruta al archivo
    :param url: url original del documento (opcional)
    :return: dict con 'text', 'metadata' o 'error'
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in [".html", ".htm"]:
        try:
            with open(path, encoding="utf-8") as f:
                content = f.read()
            return parse_html(content, url=url)
        except Exception as e:
            return {"error": str(e), "text": None, "metadata": None}
    elif ext == ".docx":
        return parse_docx(path, url=url)
    elif ext == ".xlsx":
        return parse_xlsx(path, url=url)
    elif ext == ".pdf":
        return parse_pdf(path, url=url)
    else:
        return {"error": f"Formato no soportado: {ext}", "text": None, "metadata": None}
